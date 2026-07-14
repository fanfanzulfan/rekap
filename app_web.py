import streamlit as st
import sqlite3
import pandas as pd
from fpdf import FPDF
import os
import re
import openpyxl
from openpyxl.drawing.image import Image as xlImage
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta

# Set Konfigurasi Halaman Web
st.set_page_config(page_title="Laporan & Absensi PGNCOM", page_icon="📝", layout="wide")

# ==========================================
# 1. KONFIGURASI DATABASE & FOLDER
# ==========================================
os.makedirs("temp_uploads", exist_ok=True)

def init_db():
    conn = sqlite3.connect("laporan_lembur_web.db")
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS detail_lembur (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama_pekerja TEXT,
            nik_pekerja TEXT,
            jabatan_pekerja TEXT,
            fungsi TEXT,
            nama_atasan TEXT,
            jabatan_atasan TEXT,
            nama_mengetahui TEXT,
            jabatan_mengetahui TEXT,
            satuan_kerja TEXT,
            lokasi_kerja TEXT,
            periode TEXT,
            tanggal TEXT,
            jam_mulai TEXT,
            jam_selesai TEXT,
            jenis_lembur TEXT,
            uraian_kerja TEXT,
            hasil_kerja TEXT,
            keterangan TEXT,
            evidence_path TEXT,
            logo_path TEXT
        )
    ''')
    conn.commit()
    return conn

# Helper Menyimpan File Upload dari Browser
def save_uploaded_file(uploadedfile):
    if uploadedfile is not None:
        file_path = os.path.join("temp_uploads", uploadedfile.name)
        with open(file_path, "wb") as f:
            f.write(uploadedfile.getbuffer())
        return file_path
    return ""

# ==========================================
# 2. HELPER FUNGSI (TANGGAL, DURASI, KOP SURAT)
# ==========================================
def hitung_durasi(jam_mulai, jam_selesai):
    try:
        fmt = "%H:%M"
        t1 = datetime.strptime(jam_mulai, fmt)
        t2 = datetime.strptime(jam_selesai, fmt)
        if t2 < t1: t2 += timedelta(days=1)
        tdelta = t2 - t1
        total_seconds = int(tdelta.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:00", total_seconds
    except:
        return "00:00:00", 0

def parse_indo_date(date_str):
    bulan_map = {"januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6, 
                 "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
                 "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "agu": 8, "sep": 9, "okt": 10, "nov": 11, "des": 12}
    try:
        parts = date_str.lower().strip().split()
        if len(parts) >= 3:
            d = int(parts[0])
            m = bulan_map.get(parts[1], 1)
            y = int(parts[2])
            return datetime(y, m, d)
    except: pass
    return None

def buat_kop_surat(ws, logo_path):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[1].height = 25; ws.row_dimensions[2].height = 25
    ws.merge_cells('A1:B2')
    if logo_path and os.path.exists(logo_path):
        try:
            img = xlImage(logo_path); img.width, img.height = 120, 50; ws.add_image(img, 'A1')
        except: pass
    for row in range(1, 3):
        for col in range(1, 3): ws.cell(row=row, column=col).border = thin_border
    ws.merge_cells('C1:D1'); ws.merge_cells('C2:D2')
    ws['C1'] = "PT PGAS Telekomunikasi Nusantara"; ws['C1'].alignment = Alignment(horizontal='center', vertical='center'); ws['C1'].font = Font(size=11)
    ws['C2'] = "SURAT PERINTAH PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA"; ws['C2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); ws['C2'].font = Font(size=11, bold=True)
    for row in range(1, 3):
        for col in range(3, 5): ws.cell(row=row, column=col).border = thin_border
    ws['E1'] = "FS No : 09"; ws['E1'].font = Font(size=10); ws['E1'].border = thin_border; ws['E1'].alignment = Alignment(vertical='center')
    ws['E2'] = "No Revisi : 01"; ws['E2'].font = Font(size=10); ws['E2'].border = thin_border; ws['E2'].alignment = Alignment(vertical='center')

# ==========================================
# 3. FUNGSI EXPORT KE EXCEL & PDF
# ==========================================
def export_to_excel():
    conn = sqlite3.connect("laporan_lembur_web.db")
    df_detail = pd.read_sql_query("SELECT * FROM detail_lembur", conn)
    if df_detail.empty: return None

    row_data_latest = df_detail.iloc[-1]
    safe_nama = re.sub(r'[\\/*?:"<>|]', "", str(row_data_latest['nama_pekerja']).strip())
    safe_periode = re.sub(r'[\\/*?:"<>|]', "", str(row_data_latest['periode']).strip())
    nama_file = f"temp_uploads/Laporan_Lembur_{safe_nama}_{safe_periode}.xlsx"
    
    wb = openpyxl.Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- SHEET 1: FORM ABSEN ---
    ws_absen = wb.active; ws_absen.title = "Form Absen"
    col_widths = {'B': 18, 'C': 12, 'D': 10, 'E': 10, 'F': 18, 'G': 6, 'H': 4, 'I': 4, 'J': 4, 'K': 4, 'L': 4, 'M': 4, 'N': 4, 'P': 45}
    for col, width in col_widths.items(): ws_absen.column_dimensions[col].width = width

    ws_absen.row_dimensions[1].height = 25; ws_absen.row_dimensions[2].height = 25
    ws_absen.merge_cells('B1:D2')
    if row_data_latest['logo_path'] and os.path.exists(row_data_latest['logo_path']):
        try: img = xlImage(row_data_latest['logo_path']); img.width, img.height = 120, 45; ws_absen.add_image(img, 'B1')
        except: pass
        
    ws_absen.merge_cells('E1:P1'); ws_absen['E1'] = "ABSENSI HARIAN"; ws_absen['E1'].font = Font(bold=True, size=16); ws_absen['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_absen.merge_cells('E2:P2'); ws_absen['E2'] = "PEKERJA PROJEK JDO-PGNCOM"; ws_absen['E2'].font = Font(bold=True, size=14); ws_absen['E2'].alignment = Alignment(horizontal='center', vertical='center')

    for r in range(1, 3):
        for c in range(2, 17): ws_absen.cell(row=r, column=c).border = thin_border
    
    ws_absen['B4'] = "NAMA"; ws_absen['C4'] = f": {row_data_latest['nama_pekerja']}"
    ws_absen['M4'] = "JABATAN"; ws_absen['P4'] = f": {row_data_latest['jabatan_pekerja']}"
    ws_absen['B5'] = "NIK"; ws_absen['C5'] = f": {row_data_latest['nik_pekerja']}"
    ws_absen['M5'] = "FUNGSI"; ws_absen['P5'] = f": {row_data_latest['fungsi']}"
    ws_absen['B6'] = "BULAN"; ws_absen['C6'] = f": {row_data_latest['periode']}"
    ws_absen['M6'] = "LOKASI KERJA"; ws_absen['P6'] = f": {row_data_latest['lokasi_kerja']}"

    headers = [('B10', 'B11', 'TANGGAL'), ('C10', 'C11', 'HARI'), ('D10', 'E10', 'J A M'), ('F10', 'F11', 'PARAF KARYAWAN'), ('G10', 'G11', 'W/H'), ('H10', 'N10', 'T I D A K   H A D I R'), ('P10', 'P11', 'K E T E R A N G A N')]
    for start, end, text in headers:
        ws_absen.merge_cells(f"{start}:{end}"); ws_absen[start] = text; ws_absen[start].alignment = Alignment(horizontal='center', vertical='center'); ws_absen[start].font = Font(bold=True)
    sub_headers = {'D11': 'DATANG', 'E11': 'PULANG', 'H11': 'I', 'I11': 'S', 'J11': 'C', 'K11': 'TL', 'L11': 'TG', 'M11': 'MR', 'N11': 'L'}
    for cell, text in sub_headers.items(): ws_absen[cell] = text; ws_absen[cell].alignment = Alignment(horizontal='center', vertical='center'); ws_absen[cell].font = Font(bold=True)

    for r in range(10, 12):
        for c in range(2, 17): 
            if c != 15: ws_absen.cell(row=r, column=c).border = thin_border

    start_date, end_date = None, None
    try:
        parts = str(row_data_latest['periode']).split('-')
        if len(parts) == 2:
            start_date = parse_indo_date(parts[0])
            end_date = parse_indo_date(parts[1])
    except: pass

    lembur_dict = {parse_indo_date(r['tanggal']).strftime("%Y-%m-%d"): r for _, r in df_detail.iterrows() if parse_indo_date(r['tanggal'])}
    current_row = 12; total_hari_kerja = 0
    indo_days = {0: 'Senin', 1: 'Selasa', 2: 'Rabu', 3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'}
    indo_months = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}

    if start_date and end_date:
        curr_d = start_date
        while curr_d <= end_date:
            tgl_str = f"{curr_d.day:02d} {indo_months[curr_d.month]} {curr_d.year}"
            hari_str = indo_days[curr_d.weekday()]
            dt_key = curr_d.strftime("%Y-%m-%d")
            
            jam_in, jam_out, mark_l, ket = "", "", "", ""
            if hari_str in ['Sabtu', 'Minggu']:
                if dt_key in lembur_dict:
                    lembur = lembur_dict[dt_key]
                    jam_in, jam_out = lembur['jam_mulai'], lembur['jam_selesai']
                    mark_l, ket, total_hari_kerja = "X", lembur['uraian_kerja'], total_hari_kerja + 1  
                else: ket = "Hari libur"
            else:
                total_hari_kerja += 1; jam_in = "07:00"
                if dt_key in lembur_dict:
                    lembur = lembur_dict[dt_key]
                    jam_out, mark_l, ket = lembur['jam_selesai'], "X", lembur['uraian_kerja']
                else: jam_out = "16:00"

            ws_absen.cell(row=current_row, column=2, value=tgl_str).alignment = Alignment(horizontal='center')
            ws_absen.cell(row=current_row, column=3, value=hari_str).alignment = Alignment(horizontal='center')
            ws_absen.cell(row=current_row, column=4, value=jam_in).alignment = Alignment(horizontal='center')
            ws_absen.cell(row=current_row, column=5, value=jam_out).alignment = Alignment(horizontal='center')
            ws_absen.cell(row=current_row, column=14, value=mark_l).alignment = Alignment(horizontal='center') 
            ws_absen.cell(row=current_row, column=16, value=ket).alignment = Alignment(wrap_text=True, vertical='center') 

            for c in range(2, 17):
                if c != 15: ws_absen.cell(row=current_row, column=c).border = thin_border
            current_row += 1; curr_d += timedelta(days=1)
    else: ws_absen.cell(row=current_row, column=2, value="Gagal membaca format periode.").alignment = Alignment(horizontal='center')

    # Footer Absensi
    r_base = current_row + 1
    ws_absen.cell(row=r_base, column=2, value="CATATAN (TOTAL) :").font = Font(bold=True)
    ws_absen.cell(row=r_base+1, column=2, value="Hari Kerja :"); ws_absen.cell(row=r_base+1, column=3, value=total_hari_kerja).alignment = Alignment(horizontal='center')
    ws_absen.cell(row=r_base+1, column=3).border = Border(bottom=Side(style='thin')); ws_absen.cell(row=r_base+1, column=4).border = Border(bottom=Side(style='thin')); ws_absen.cell(row=r_base+1, column=5, value="hari.")
    ws_absen.cell(row=r_base+3, column=2, value="KETERANGAN :").font = Font(bold=True)
    ws_absen.merge_cells(start_row=r_base+4, start_column=8, end_row=r_base+4, end_column=12); ws_absen.cell(row=r_base+4, column=8, value="Mengetahui,").alignment = Alignment(horizontal='center')
    ws_absen.cell(row=r_base+4, column=16, value="Dibuat oleh").alignment = Alignment(horizontal='center')
    ws_absen.cell(row=r_base+5, column=2, value="I").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+5, column=3, value=": Ijin")
    ws_absen.cell(row=r_base+5, column=6, value="TL").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+5, column=7, value=": Terlambat")
    ws_absen.cell(row=r_base+6, column=2, value="S").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+6, column=3, value=": Sakit")
    ws_absen.cell(row=r_base+6, column=6, value="TG").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+6, column=7, value=": Tugas")
    ws_absen.cell(row=r_base+7, column=2, value="C").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+7, column=3, value=": Cuti")
    ws_absen.cell(row=r_base+7, column=6, value="MR").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+7, column=7, value=": Mangkir")
    ws_absen.cell(row=r_base+9, column=2, value="L").alignment = Alignment(horizontal='center'); ws_absen.cell(row=r_base+9, column=3, value=": Lembur")

    ws_absen.merge_cells(start_row=r_base+8, start_column=8, end_row=r_base+8, end_column=12)
    ws_absen.cell(row=r_base+8, column=8, value=f"({row_data_latest['nama_atasan']})").font = Font(bold=True); ws_absen.cell(row=r_base+8, column=8).alignment = Alignment(horizontal='center')
    ws_absen.cell(row=r_base+8, column=16, value=f"({row_data_latest['nama_pekerja']})").font = Font(bold=True); ws_absen.cell(row=r_base+8, column=16).alignment = Alignment(horizontal='center')
    ws_absen.merge_cells(start_row=r_base+9, start_column=8, end_row=r_base+9, end_column=12)
    ws_absen.cell(row=r_base+9, column=8, value=f"({row_data_latest['jabatan_atasan']})").font = Font(bold=True); ws_absen.cell(row=r_base+9, column=8).alignment = Alignment(horizontal='center', wrap_text=True)
    ws_absen.cell(row=r_base+9, column=16, value=f"({row_data_latest['jabatan_pekerja']})").font = Font(bold=True); ws_absen.cell(row=r_base+9, column=16).alignment = Alignment(horizontal='center', wrap_text=True)

    # --- SHEET 2: REKAP LEMBUR ---
    ws_rekap = wb.create_sheet(title="Rekap lembur")
    lebar_kolom_rekap = {'A': 5, 'B': 40, 'C': 15, 'D': 12, 'E': 12, 'F': 22, 'G': 25}
    for col, width in lebar_kolom_rekap.items(): ws_rekap.column_dimensions[col].width = width

    ws_rekap.append(["PT PGAS Telekomunikasi Nusantara"]); ws_rekap.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws_rekap.append(["SURAT PERINTAH PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA"]); ws_rekap.cell(row=2, column=1).font = Font(bold=True, size=12)
    ws_rekap.append([]); ws_rekap.append(["Nama", ":", row_data_latest['nama_pekerja']])
    ws_rekap.append(["Jabatan", ":", row_data_latest['jabatan_pekerja']]); ws_rekap.append(["Satuan Kerja", ":", row_data_latest['satuan_kerja']])
    ws_rekap.append(["Periode", ":", row_data_latest['periode']]); ws_rekap.append([])
    
    headers = ["No", "Kerja Tambahan", "Tgl", "Jam Mulai", "Jam Selesai", "Lama Kerja Tambahan", "Keterangan"]
    ws_rekap.append(headers)
    header_row = ws_rekap.max_row
    for col in range(1, 8): cell = ws_rekap.cell(row=header_row, column=col); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border

    total_detik = 0
    for index, row in df_detail.iterrows():
        durasi_str, detik = hitung_durasi(row['jam_mulai'], row['jam_selesai'])
        total_detik += detik
        ws_rekap.append([index + 1, row['uraian_kerja'], row['tanggal'], row['jam_mulai'], row['jam_selesai'], durasi_str, row['keterangan']])
        r_idx = ws_rekap.max_row
        for col in range(1, 8):
            cell = ws_rekap.cell(row=r_idx, column=col)
            cell.border = thin_border
            if col not in [2, 7]: cell.alignment = Alignment(horizontal='center', vertical='center')
            else: cell.alignment = Alignment(vertical='center', wrap_text=True)

    t_h, rem = divmod(total_detik, 3600); t_m, t_s = divmod(rem, 60)
    ws_rekap.append(["", "Jumlah", "", "", "", f"{t_h:02d}:{t_m:02d}:00", ""])
    last_row = ws_rekap.max_row
    ws_rekap.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=5)
    ws_rekap.cell(row=last_row, column=2).alignment = Alignment(horizontal='right', vertical='center'); ws_rekap.cell(row=last_row, column=2).font = Font(bold=True)
    ws_rekap.cell(row=last_row, column=6).font = Font(bold=True); ws_rekap.cell(row=last_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 8): ws_rekap.cell(row=last_row, column=col).border = thin_border
    
    ws_rekap.append([]); ws_rekap.append([])
    ttd_row = ws_rekap.max_row
    ws_rekap.cell(row=ttd_row, column=2, value="Mengetahui,").alignment = Alignment(horizontal='center'); ws_rekap.cell(row=ttd_row, column=6, value="Menyetujui,").alignment = Alignment(horizontal='center')
    nama_row = ttd_row + 4
    ws_rekap.cell(row=nama_row, column=2, value=row_data_latest['nama_mengetahui']).font = Font(bold=True, underline="single"); ws_rekap.cell(row=nama_row, column=2).alignment = Alignment(horizontal='center')
    ws_rekap.cell(row=nama_row+1, column=2, value=row_data_latest['jabatan_mengetahui']).alignment = Alignment(horizontal='center')
    ws_rekap.cell(row=nama_row, column=6, value=row_data_latest['nama_atasan']).font = Font(bold=True, underline="single"); ws_rekap.cell(row=nama_row, column=6).alignment = Alignment(horizontal='center')
    ws_rekap.cell(row=nama_row+1, column=6, value=row_data_latest['jabatan_atasan']).alignment = Alignment(horizontal='center')

    # --- SHEET 3+: HARIAN & EVIDENCE ---
    for index, row in df_detail.iterrows():
        tanggal_safe = str(row['tanggal']).replace('/', '-')
        ws = wb.create_sheet(title=f"{tanggal_safe[:31]}")
        ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 3; ws.column_dimensions['D'].width = 45; ws.column_dimensions['E'].width = 20
        buat_kop_surat(ws, row['logo_path'])
        
        biasa_chk = "[ v ]" if row['jenis_lembur'] == "Biasa" else "[   ]"
        emrg_chk = "[ v ]" if row['jenis_lembur'] == "Emergency" else "[   ]"
        teks_jenis_lembur = f"Biasa {biasa_chk}       Emergency {emrg_chk}"
        
        ws.append([]); ws.append(["A. SURAT PERINTAH PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True); ws.append([])
        ws.append(["Satuan Kerja", "", ":", row['satuan_kerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Tanggal", "", ":", row['tanggal']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append([]); ws.append(["Kepada pekerja tersebut di bawah ini, ditugaskan untuk Kerja Tambahan"])
        ws.append(["Nama", "", ":", row['nama_pekerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Jenis Kerja Lembur", "", ":", teks_jenis_lembur]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Uraian Kerja Lembur", "", ":", row['uraian_kerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2); ws.append([])

        row_ttd_a = ws.max_row + 1
        ws.cell(row=row_ttd_a, column=1, value="Yang melaksanakan Kerja Tambahan,").alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_ttd_a, start_column=1, end_row=row_ttd_a, end_column=2)
        ws.cell(row=row_ttd_a, column=4, value="Menyetujui,").alignment = Alignment(horizontal='center')
        row_nama_a = row_ttd_a + 4
        ws.cell(row=row_nama_a, column=1, value=row['nama_pekerja']).font = Font(bold=True, underline="single"); ws.cell(row=row_nama_a, column=1).alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_nama_a, start_column=1, end_row=row_nama_a, end_column=2)
        ws.cell(row=row_nama_a, column=4, value=row['nama_atasan']).font = Font(bold=True, underline="single"); ws.cell(row=row_nama_a, column=4).alignment = Alignment(horizontal='center')
        row_jab_a = row_nama_a + 1
        ws.cell(row=row_jab_a, column=1, value=row['jabatan_pekerja']).alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_jab_a, start_column=1, end_row=row_jab_a, end_column=2)
        ws.cell(row=row_jab_a, column=4, value=row['jabatan_atasan']).alignment = Alignment(horizontal='center')
        
        ws.append([]); ws.append([])
        
        ws.append(["B. SURAT PERTANGGUNGJAWABAN PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Perintah Kerja Tambahan pada:"])
        ws.append(["Tanggal", "", ":", row['tanggal']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Jam", "", ":", f"{row['jam_mulai']} s/d {row['jam_selesai']}"]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Jenis Kerja Tambahan", "", ":", teks_jenis_lembur]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Hasil Kerja Tambahan", "", ":", row['hasil_kerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append([]); ws.append(["Telah dilaksanakan oleh:"])
        ws.append(["Nama", "", ":", row['nama_pekerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.append(["Satuan Kerja", "", ":", row['satuan_kerja']]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2); ws.append([])

        row_ttd_b = ws.max_row + 1
        ws.cell(row=row_ttd_b, column=1, value="Yang melaksanakan Kerja Tambahan,").alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_ttd_b, start_column=1, end_row=row_ttd_b, end_column=2)
        ws.cell(row=row_ttd_b, column=4, value="Menyetujui,").alignment = Alignment(horizontal='center')
        row_nama_b = row_ttd_b + 4
        ws.cell(row=row_nama_b, column=1, value=row['nama_pekerja']).font = Font(bold=True, underline="single"); ws.cell(row=row_nama_b, column=1).alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_nama_b, start_column=1, end_row=row_nama_b, end_column=2)
        ws.cell(row=row_nama_b, column=4, value=row['nama_atasan']).font = Font(bold=True, underline="single"); ws.cell(row=row_nama_b, column=4).alignment = Alignment(horizontal='center')
        row_jab_b = row_nama_b + 1
        ws.cell(row=row_jab_b, column=1, value=row['jabatan_pekerja']).alignment = Alignment(horizontal='center'); ws.merge_cells(start_row=row_jab_b, start_column=1, end_row=row_jab_b, end_column=2)
        ws.cell(row=row_jab_b, column=4, value=row['jabatan_atasan']).alignment = Alignment(horizontal='center')

        ws_evd = wb.create_sheet(title=f"Evd {tanggal_safe[:20]}")
        buat_kop_surat(ws_evd, row['logo_path']) 
        ws_evd.append([]); ws_evd.append([f"Bukti Pekerjaan Tanggal: {row['tanggal']}"]); ws_evd.cell(row=ws_evd.max_row, column=1).font = Font(bold=True)
        if row['evidence_path'] and os.path.exists(row['evidence_path']):
            try: img = xlImage(row['evidence_path']); img.width, img.height = 400, 400; ws_evd.add_image(img, 'A6')
            except: pass
                
    wb.save(nama_file)
    conn.close()
    return nama_file

def export_to_pdf():
    conn = sqlite3.connect("laporan_lembur_web.db")
    df_detail = pd.read_sql_query("SELECT * FROM detail_lembur", conn)
    if df_detail.empty: return None

    row_data_latest = df_detail.iloc[-1]
    safe_nama = re.sub(r'[\\/*?:"<>|]', "", str(row_data_latest['nama_pekerja']).strip())
    safe_periode = re.sub(r'[\\/*?:"<>|]', "", str(row_data_latest['periode']).strip())
    nama_file = f"temp_uploads/Laporan_Lembur_{safe_nama}_{safe_periode}.pdf"
    pdf = FPDF()
    
    def buat_kop_surat_pdf(pdf_obj, logo_path):
        if logo_path and os.path.exists(logo_path): pdf_obj.image(logo_path, x=10, y=10, w=35)
        pdf_obj.set_font("Arial", 'B', 12); pdf_obj.cell(0, 8, "PT PGAS Telekomunikasi Nusantara", ln=True, align='C')
        pdf_obj.set_font("Arial", 'B', 11); pdf_obj.cell(0, 6, "SURAT PERINTAH PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA", ln=True, align='C')
        pdf_obj.line(10, 25, 200, 25); pdf_obj.ln(10)

    pdf.add_page(); buat_kop_surat_pdf(pdf, row_data_latest['logo_path'])
    pdf.set_font("Arial", '', 10)
    pdf.cell(30, 6, "Nama", 0, 0); pdf.cell(0, 6, f": {row_data_latest['nama_pekerja']}", 0, 1)
    pdf.cell(30, 6, "Jabatan", 0, 0); pdf.cell(0, 6, f": {row_data_latest['jabatan_pekerja']}", 0, 1)
    pdf.cell(30, 6, "Satuan Kerja", 0, 0); pdf.cell(0, 6, f": {row_data_latest['satuan_kerja']}", 0, 1)
    pdf.cell(30, 6, "Periode", 0, 0); pdf.cell(0, 6, f": {row_data_latest['periode']}", 0, 1); pdf.ln(5)

    pdf.set_font("Arial", 'B', 9); col_widths = [10, 60, 25, 20, 20, 25, 30]
    headers = ["No", "Kerja Tambahan", "Tgl", "Jam Mulai", "Selesai", "Durasi", "Keterangan"]
    for i in range(len(headers)): pdf.cell(col_widths[i], 8, headers[i], border=1, align='C')
    pdf.ln()

    pdf.set_font("Arial", '', 8); total_detik = 0
    for index, row in df_detail.iterrows():
        durasi_str, detik = hitung_durasi(row['jam_mulai'], row['jam_selesai']); total_detik += detik
        uraian = (row['uraian_kerja'][:35] + '..') if len(row['uraian_kerja']) > 35 else row['uraian_kerja']
        ket = (row['keterangan'][:20] + '..') if len(row['keterangan']) > 20 else row['keterangan']
        pdf.cell(col_widths[0], 6, str(index + 1), border=1, align='C'); pdf.cell(col_widths[1], 6, uraian, border=1, align='L')
        pdf.cell(col_widths[2], 6, str(row['tanggal']), border=1, align='C'); pdf.cell(col_widths[3], 6, str(row['jam_mulai']), border=1, align='C')
        pdf.cell(col_widths[4], 6, str(row['jam_selesai']), border=1, align='C'); pdf.cell(col_widths[5], 6, durasi_str, border=1, align='C')
        pdf.cell(col_widths[6], 6, ket, border=1, align='C'); pdf.ln()

    t_h, rem = divmod(total_detik, 3600); t_m, t_s = divmod(rem, 60); pdf.set_font("Arial", 'B', 9)
    pdf.cell(sum(col_widths[:5]), 8, "Jumlah", border=1, align='R'); pdf.cell(col_widths[5], 8, f"{t_h:02d}:{t_m:02d}:00", border=1, align='C'); pdf.cell(col_widths[6], 8, "", border=1, align='C'); pdf.ln(15)

    pdf.set_font("Arial", '', 10); pdf.cell(95, 6, "Mengetahui,", 0, 0, 'C'); pdf.cell(95, 6, "Menyetujui,", 0, 1, 'C'); pdf.ln(15)
    pdf.set_font("Arial", 'BU', 10); pdf.cell(95, 6, row_data_latest['nama_mengetahui'], 0, 0, 'C'); pdf.cell(95, 6, row_data_latest['nama_atasan'], 0, 1, 'C')
    pdf.set_font("Arial", '', 10); pdf.cell(95, 6, row_data_latest['jabatan_mengetahui'], 0, 0, 'C'); pdf.cell(95, 6, row_data_latest['jabatan_atasan'], 0, 1, 'C')

    for index, row in df_detail.iterrows():
        pdf.add_page(); buat_kop_surat_pdf(pdf, row['logo_path'])
        pdf.set_font("Arial", 'B', 10); pdf.cell(0, 8, "A. SURAT PERINTAH PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA", ln=True)
        pdf.set_font("Arial", '', 10)
        pdf.cell(40, 6, "Satuan Kerja", 0, 0); pdf.cell(0, 6, f": {row['satuan_kerja']}", 0, 1)
        pdf.cell(40, 6, "Tanggal", 0, 0); pdf.cell(0, 6, f": {row['tanggal']}", 0, 1); pdf.ln(3)
        pdf.cell(0, 6, "Kepada pekerja tersebut di bawah ini, ditugaskan untuk Kerja Tambahan", ln=True)
        pdf.cell(40, 6, "Nama", 0, 0); pdf.cell(0, 6, f": {row['nama_pekerja']}", 0, 1)
        biasa_chk = "[ v ]" if row['jenis_lembur'] == "Biasa" else "[   ]"; emrg_chk = "[ v ]" if row['jenis_lembur'] == "Emergency" else "[   ]"
        pdf.cell(40, 6, "Jenis Kerja Lembur", 0, 0); pdf.cell(0, 6, f": {biasa_chk} Biasa     {emrg_chk} Emergency", 0, 1)
        pdf.cell(40, 6, "Uraian Kerja Lembur", 0, 0); pdf.cell(5, 6, ":", 0, 0); pdf.multi_cell(0, 6, row['uraian_kerja']); pdf.ln(10)

        pdf.cell(100, 6, "Yang melaksanakan Kerja Tambahan,", 0, 0, 'C'); pdf.cell(90, 6, "Menyetujui,", 0, 1, 'C'); pdf.ln(15)
        pdf.set_font("Arial", 'BU', 10); pdf.cell(100, 6, row['nama_pekerja'], 0, 0, 'C'); pdf.cell(90, 6, row['nama_atasan'], 0, 1, 'C')
        pdf.set_font("Arial", '', 10); pdf.cell(100, 6, row['jabatan_pekerja'], 0, 0, 'C'); pdf.cell(90, 6, row['jabatan_atasan'], 0, 1, 'C'); pdf.ln(5); pdf.line(10, pdf.get_y(), 200, pdf.get_y()); pdf.ln(5)

        pdf.set_font("Arial", 'B', 10); pdf.cell(0, 8, "B. SURAT PERTANGGUNGJAWABAN PELAKSANAAN PEKERJAAN DI LUAR WAKTU KERJA", ln=True)
        pdf.set_font("Arial", '', 10); pdf.cell(0, 6, "Perintah Kerja Tambahan pada:", ln=True)
        pdf.cell(40, 6, "Tanggal", 0, 0); pdf.cell(0, 6, f": {row['tanggal']}", 0, 1)
        pdf.cell(40, 6, "Jam", 0, 0); pdf.cell(0, 6, f": {row['jam_mulai']} s/d {row['jam_selesai']}", 0, 1)
        pdf.cell(40, 6, "Jenis Kerja Tambahan", 0, 0); pdf.cell(0, 6, f": {biasa_chk} Biasa     {emrg_chk} Emergency", 0, 1)
        pdf.cell(40, 6, "Hasil Kerja Tambahan", 0, 0); pdf.cell(5, 6, ":", 0, 0); pdf.multi_cell(0, 6, row['hasil_kerja']); pdf.ln(3)
        pdf.cell(0, 6, "Telah dilaksanakan oleh:", ln=True)
        pdf.cell(40, 6, "Nama", 0, 0); pdf.cell(0, 6, f": {row['nama_pekerja']}", 0, 1)
        pdf.cell(40, 6, "Satuan Kerja", 0, 0); pdf.cell(0, 6, f": {row['satuan_kerja']}", 0, 1); pdf.ln(10)

        pdf.cell(100, 6, "Yang melaksanakan Kerja Tambahan,", 0, 0, 'C'); pdf.cell(90, 6, "Menyetujui,", 0, 1, 'C'); pdf.ln(15)
        pdf.set_font("Arial", 'BU', 10); pdf.cell(100, 6, row['nama_pekerja'], 0, 0, 'C'); pdf.cell(90, 6, row['nama_atasan'], 0, 1, 'C')
        pdf.set_font("Arial", '', 10); pdf.cell(100, 6, row['jabatan_pekerja'], 0, 0, 'C'); pdf.cell(90, 6, row['jabatan_atasan'], 0, 1, 'C')
        
        if row['evidence_path'] and os.path.exists(row['evidence_path']):
            pdf.add_page(); buat_kop_surat_pdf(pdf, row['logo_path']) 
            pdf.set_font("Arial", 'B', 12); pdf.cell(0, 10, f"Evidence Pelaksanaan Pekerjaan - {row['tanggal']}", ln=True, align='C'); pdf.ln(5)
            try: pdf.image(row['evidence_path'], x=20, w=170)
            except: pass
    pdf.output(nama_file)
    conn.close()
    return nama_file


# ==========================================
# 4. ANTARMUKA WEB (STREAMLIT)
# ==========================================
def main():
    st.title("📄 Aplikasi Laporan & Absensi PGNCOM")
    st.markdown("Aplikasi web untuk mencatat laporan harian lembur dan meng-generate file rekap Excel beserta PDF secara otomatis.")

    conn = init_db()

    # --- BAGIAN 1: FORM DATA PERSONIL ---
    with st.expander("👤 DATA PERSONIL & PERIODE (Wajib Diisi)", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            nama = st.text_input("Nama Pekerja (Kosongkan/Ganti sesuai data asli):", value="")
            nik = st.text_input("NIK Pekerja:", value="")
            jabatan = st.text_input("Jabatan Pekerja:", value="")
            satker = st.text_input("Satuan Kerja:", value="Jasa Dukungan Operasional")
            lokasi = st.text_input("Lokasi Kerja:", value="Grha Pertamina")
            fungsi = st.text_input("Fungsi:", value="SS ICT - HQFS")
            logo_file = st.file_uploader("Upload Logo PGN (png/jpg):", type=['png', 'jpg', 'jpeg'])

        with col2:
            periode = st.text_input("Bulan/Periode Absen:", value="14 Juni 2026 - 13 Juli 2026")
            st.markdown("---")
            nama_atasan = st.text_input("Nama Penyetuju:", value="Erna Danawianti")
            jabatan_atasan = st.text_input("Jabatan Penyetuju:", value="Officer II HQ Field Service")
            nama_mengetahui = st.text_input("Nama Mengetahui:", value="Chyntia Rahmi Lubis")
            jabatan_mengetahui = st.text_input("Jabatan Mengetahui:", value="Jr. Analyst, Resource Planning")

    # --- BAGIAN 2: FORM LEMBUR HARIAN ---
    with st.expander("📅 INPUT DATA LEMBUR HARIAN", expanded=True):
        col3, col4 = st.columns(2)
        with col3:
            tgl = st.text_input("Tanggal (ex: 15 Juni 2026):")
            jam_mulai = st.text_input("Jam Mulai (ex: 18:00):")
            jam_selesai = st.text_input("Jam Selesai (ex: 00:00):")
            jenis = st.radio("Jenis Kerja Lembur:", ["Emergency", "Biasa"], horizontal=True)
            ket = st.text_input("Keterangan (ex: Request Pak Diki):")
            
        with col4:
            uraian = st.text_area("Uraian Kerja (Perintah):")
            hasil = st.text_area("Hasil Kerja (Pertanggungjawaban):")
            evd_file = st.file_uploader("Upload Bukti Pekerjaan (png/jpg):", type=['png', 'jpg', 'jpeg'])

        if st.button("💾 Simpan Data Harian ke Database", type="primary", use_container_width=True):
            if not nama or not tgl:
                st.error("Gagal Menyimpan! Harap isi Nama Pekerja dan Tanggal terlebih dahulu.")
            else:
                logo_path = save_uploaded_file(logo_file)
                evd_path = save_uploaded_file(evd_file)
                
                c = conn.cursor()
                c.execute('''
                    INSERT INTO detail_lembur (nama_pekerja, nik_pekerja, jabatan_pekerja, fungsi, nama_atasan, jabatan_atasan, 
                    nama_mengetahui, jabatan_mengetahui, satuan_kerja, lokasi_kerja, periode,
                    tanggal, jam_mulai, jam_selesai, jenis_lembur, uraian_kerja, hasil_kerja, keterangan, evidence_path, logo_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (nama, nik, jabatan, fungsi, nama_atasan, jabatan_atasan, nama_mengetahui, jabatan_mengetahui, satker, lokasi, periode,
                      tgl, jam_mulai, jam_selesai, jenis, uraian, hasil, ket, evd_path, logo_path))
                conn.commit()
                st.success(f"Data harian tanggal {tgl} atas nama {nama} berhasil disimpan!")

    # --- BAGIAN 3: EXPORT AREA ---
    st.markdown("### 📥 Export Laporan")
    st.info("Pastikan Anda sudah menyimpan semua data harian sebelum menekan tombol Generate di bawah ini.")
    
    col5, col6, col7 = st.columns(3)
    
    with col5:
        if st.button("📊 Buat File EXCEL"):
            excel_path = export_to_excel()
            if excel_path:
                with open(excel_path, "rb") as f:
                    st.download_button(label="Unduh File EXCEL Sekarang", data=f, file_name=os.path.basename(excel_path), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("Data masih kosong, tidak ada yang bisa di-export.")

    with col6:
        if st.button("📑 Buat File PDF"):
            pdf_path = export_to_pdf()
            if pdf_path:
                with open(pdf_path, "rb") as f:
                    st.download_button(label="Unduh File PDF Sekarang", data=f, file_name=os.path.basename(pdf_path), mime="application/pdf")
            else:
                st.warning("Data masih kosong, tidak ada yang bisa di-export.")
                
    with col7:
        if st.button("🗑️ Reset Database (Bulan Baru)"):
            c = conn.cursor()
            c.execute("DELETE FROM detail_lembur")
            conn.commit()
            st.success("Database berhasil dikosongkan. Siap untuk input bulan baru!")

if __name__ == "__main__":
    main()